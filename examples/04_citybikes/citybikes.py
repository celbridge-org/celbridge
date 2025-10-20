import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


API_URL = "https://api.citybik.es/v2/networks/dublinbikes"
OUTPUT_XLSX = "04_citybikes/citybikes.xlsx"

def main():

    # Download bike station data
    response = requests.get(API_URL, timeout=15)
    response.raise_for_status()
    stations = response.json()["network"]["stations"]

    # Create workbook & sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Bikes"

    # Write the header row
    headers = ["Station Name", "Available Bikes", "Empty Slots"]
    ws.append(headers)

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15

    # Apply header style (bold, shaded, centered)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Sort stations by name
    stations.sort(key=lambda s: s.get("name", "").lower())

    # Write a row for each station
    for s in stations:
        name = s.get("name", "")
        free = s.get("free_bikes") or 0
        empty = s.get("empty_slots") or 0
        ws.append([name, free, empty])

    # Save the Excel file
    wb.save(OUTPUT_XLSX)
    print(f"Saved {OUTPUT_XLSX} with {len(stations)} stations.")

if __name__ == "__main__":
    main()
